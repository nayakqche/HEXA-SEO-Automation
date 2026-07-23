"""
Pipeline — orchestrates a full SEO run.

  primary URLs + secondary URLs  →  grounding context + URL inventory
                                 →  per keyword: Claude writes JSON blog
                                 →  links validated against URL inventory
                                 →  per image block: Gemini hero + diagrams
                                 →  outputs/<run>/<post>/{json,md,html,pdf,docx,*.png}

Yields server-sent events so the web UI can show a live log.
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import os
import re
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from . import editing, image_gen, renderer, store, uploads
from .blog_writer import validate_and_clean_links, write_blog
from .scraper import GroundingContext, build_context

OUTPUT_DIR = Path("outputs")


# ── Input parsing ──────────────────────────────────────────────────────────

_HEADER_CELLS = {"keyword", "keywords", "term", "query"}


def _dedupe_clean(cells) -> list[str]:
    keywords: list[str] = []
    seen = set()
    for cell in cells:
        kw = (cell or "").strip().strip('"')
        if not kw or kw.lower() in _HEADER_CELLS:
            continue
        key = kw.lower()
        if key not in seen:
            seen.add(key)
            keywords.append(kw)
    return keywords


def _keywords_from_xlsx(data: bytes) -> list[str]:
    """Read keywords from an Excel .xlsx file (first non-empty cell of each row)."""
    import io as _io
    from openpyxl import load_workbook

    wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    cells: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is None:
                    continue
                cells.append(str(value))
                break  # only the first non-empty cell in each row
    wb.close()
    return _dedupe_clean(cells)


def parse_keywords(file_storage_or_text) -> list[str]:
    """
    Accept a CSV, an Excel .xlsx upload, or pasted text. Returns clean,
    de-duplicated keywords. Raises ValueError with a clear message for
    unsupported binary files (old .xls, .pdf, images, etc.).
    """
    if hasattr(file_storage_or_text, "read"):
        raw = file_storage_or_text.read()
    else:
        raw = str(file_storage_or_text)

    if isinstance(raw, bytes):
        # Excel .xlsx (and .docx etc.) are ZIP archives → magic bytes "PK\x03\x04".
        if raw[:4] == b"PK\x03\x04":
            try:
                return _keywords_from_xlsx(raw)
            except Exception as exc:  # noqa: BLE001
                raise ValueError(
                    f"Couldn't read that Excel file ({exc}). "
                    "Re-save it as .xlsx or export to .csv."
                ) from exc
        # Old binary .xls (BIFF) starts with this OLE2 signature.
        if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            raise ValueError(
                "That's an old-format .xls file. Please 'Save As' .xlsx or .csv "
                "and upload again."
            )
        text = raw.decode("utf-8-sig", errors="replace")
        # Guard: NUL bytes or lots of replacement chars mean binary junk.
        if "\x00" in text or text.count("�") > max(5, len(text) * 0.02):
            raise ValueError(
                "That file doesn't look like a CSV or Excel sheet. Upload a .csv "
                "or .xlsx with one keyword per row, or paste keywords as text."
            )
    else:
        text = raw

    # splitlines() handles \n, \r\n and lone \r (Excel/Mac CSVs) uniformly.
    cells = []
    for row in csv.reader(text.splitlines()):
        if row:
            cells.append(row[0])  # first column = keyword
    return _dedupe_clean(cells)


def parse_urls(text: str) -> list[str]:
    if not text:
        return []
    parts = re.split(r"[\s,]+", text.strip())
    seen, out = set(), []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if not p.startswith(("http://", "https://")):
            p = "https://" + p
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


# ── Slugs ──────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    text = re.sub(r"[^\w\s-]", "", (text or "").lower()).strip()
    return re.sub(r"[\s_-]+", "-", text)[:60] or "post"


# ── URL inventory (what Claude is allowed to link to) ──────────────────────

def _build_inventory(ctx: GroundingContext, extra_urls: list[str]) -> tuple[str, str, list[str], list[str]]:
    """
    Build the (primary_inventory_text, secondary_inventory_text,
              primary_url_list, secondary_url_list) tuple.
    """
    prim_pages = [(p.url, p.title) for p in ctx.primary]
    # Include any extra primary URLs the user passed even if we couldn't scrape them.
    for u in extra_urls:
        if u and not any(u == p[0] for p in prim_pages):
            prim_pages.append((u, urlpath_title(u)))
    sec_pages = [(p.url, p.title) for p in ctx.secondary]

    def fmt(pairs):
        if not pairs:
            return "(none)"
        return "\n".join(f"- {url}  →  \"{title[:80]}\"" for url, title in pairs)

    prim_urls = [u for u, _ in prim_pages]
    sec_urls = [u for u, _ in sec_pages]
    return fmt(prim_pages), fmt(sec_pages), prim_urls, sec_urls


def urlpath_title(url: str) -> str:
    """Derive a readable title from a URL when we couldn't fetch the page."""
    from urllib.parse import urlparse
    p = urlparse(url)
    return (p.netloc + p.path).rstrip("/") or url


# ── Main run ───────────────────────────────────────────────────────────────

def run(
    keywords: list[str],
    brand_website: str,
    *,
    primary_urls: list[str],
    secondary_urls: list[str],
    extra_instructions: str = "",
    make_images: bool = True,
    max_pages: int = 12,
    fmt: str = "paragraph",
    target_words: int = 1400,
    personalized: bool = False,
):
    OUTPUT_DIR.mkdir(exist_ok=True)
    run_id = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = OUTPUT_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    # If Pexels isn't configured, silently disable image lookup so the
    # whole pipeline doesn't fail per-keyword. The blog still produces.
    if make_images and not os.getenv("PEXELS_API_KEY"):
        make_images = False
        yield {"event": "warn", "message":
               "PEXELS_API_KEY not set — skipping stock photo lookup. "
               "Blog text + JSON/MD/PDF/DOCX will still be produced "
               "with image placeholders."}

    yield {"event": "start", "run_id": run_id, "total": len(keywords),
           "make_images": make_images}

    # 1. Build grounding context + URL inventory.
    yield {"event": "status",
           "message": f"Scraping {brand_website} + "
                      f"{len(primary_urls)} primary, {len(secondary_urls)} secondary source(s)…"}
    try:
        ctx = build_context(brand_website, primary_urls, secondary_urls, max_pages=max_pages)
    except Exception as exc:  # noqa: BLE001
        yield {"event": "error", "fatal": True, "message": str(exc)}
        return

    for note in ctx.notes:
        yield {"event": "warn", "message": note}

    prim_inv, sec_inv, allow_primary, allow_secondary = _build_inventory(
        ctx, [brand_website] + primary_urls
    )
    # Belt-and-suspenders: even if the scraper found nothing, the brand
    # website itself must be a valid internal link target so posts still
    # carry at least one Hexa link.
    if brand_website and brand_website not in allow_primary:
        allow_primary.append(brand_website)
        prim_inv = (prim_inv + "\n" if prim_inv and prim_inv != "(none)" else "") \
            + f"- {brand_website}  →  \"Hexa Climate home\""

    yield {
        "event": "grounded",
        "primary_pages": len(ctx.primary),
        "secondary_pages": len(ctx.secondary),
        "chars": ctx.char_count,
        "logo_url": ctx.logo_url,
        "fmt": fmt,
        "target_words": target_words,
        "message": (
            f"Grounded on {len(ctx.primary)} primary + {len(ctx.secondary)} secondary "
            f"page(s) ({ctx.char_count:,} chars). "
            f"Link inventory: {len(allow_primary)} internal, {len(allow_secondary)} citation. "
            f"Format: {fmt}, ~{target_words} words/post."
        ),
    }

    primary_text = ctx.primary_text()
    secondary_text = ctx.secondary_text()

    # Fold the upload container into PRIMARY context. Uploaded text/tables become
    # grounding material; uploaded images/tables get embedded into each post.
    upload_text, up_images, up_tables = uploads.resource_context()
    if upload_text:
        primary_text += "\n" + upload_text
    if up_images or up_tables or upload_text:
        yield {"event": "status",
               "message": f"Included {len(up_images)} uploaded image(s), "
                          f"{len(up_tables)} table(s), and pasted/other files as "
                          f"primary resources."}

    # PERSONALIZED MODE: the post must be ABOUT the uploaded resource, not about
    # the queue keyword. Derive the real topic from the uploads / brief and use
    # it as the SEO focus so the title, slug, and headings revolve around it.
    if personalized:
        topic = _derive_topic(up_images, up_tables, extra_instructions,
                              keywords[0] if keywords else "")
        keywords = [topic]
        yield {"event": "status",
               "message": f"Personalized post topic (from your resources): {topic}"}

    # Shared, read-only config passed to each worker.
    job = {
        "run_id": run_id, "run_dir": run_dir,
        "primary_text": primary_text, "secondary_text": secondary_text,
        "prim_inv": prim_inv, "sec_inv": sec_inv,
        "allow_primary": allow_primary, "allow_secondary": allow_secondary,
        "extra_instructions": extra_instructions, "fmt": fmt,
        "target_words": target_words, "make_images": make_images,
        "up_images": up_images, "up_tables": up_tables,
        "personalized": personalized,
        "media_brief": _media_brief(up_images, up_tables, personalized=personalized),
    }

    concurrency = max(1, min(int(os.getenv("CONCURRENCY", "1")), 8))
    yield {"event": "status",
           "message": f"Writing {len(keywords)} post(s), {concurrency} at a time…"}

    results: list[dict] = []
    done_count = 0
    # Run keywords in parallel — the work is I/O-bound (waiting on the LLM /
    # image API), so a thread pool gives a big speedup without extra CPU.
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {
            executor.submit(_process_keyword, i, kw, job): (i, kw)
            for i, kw in enumerate(keywords, start=1)
        }
        for future in as_completed(futures):
            i, kw = futures[future]
            try:
                record = future.result()
            except Exception as exc:  # noqa: BLE001 — never let one keyword kill the run
                done_count += 1
                yield {"event": "keyword_error", "index": i, "keyword": kw,
                       "stage": "blog", "message": f"{type(exc).__name__}: {exc}",
                       "done": done_count, "total": len(keywords)}
                continue
            done_count += 1
            results.append(record)
            yield {"event": "keyword_done", "done": done_count,
                   "total": len(keywords), **record}

    results.sort(key=lambda r: r["index"])
    (run_dir / "index.json").write_text(json.dumps(results, indent=2))
    yield {"event": "done", "run_id": run_id, "count": len(results),
           "message": f"Finished {len(results)} of {len(keywords)} posts."}


def _process_keyword(i: int, keyword: str, job: dict) -> dict:
    """
    Generate one blog end-to-end (write → validate links → images → all
    output formats) and return its UI record. Runs inside a worker thread.
    Raises on a fatal blog-write failure so the caller can report it.
    """
    run_id = job["run_id"]
    run_dir = job["run_dir"]

    written = write_blog(
        keyword, job["primary_text"], job["secondary_text"],
        job["prim_inv"], job["sec_inv"],
        extra_instructions=job["extra_instructions"],
        fmt=job["fmt"], target_words=job["target_words"],
        media_brief=job.get("media_brief", ""),
        primary_urls=job.get("allow_primary", []),
    )

    post = written["post"]
    slug = _slugify(post.get("slug") or post.get("meta", {}).get("title") or keyword)
    post["slug"] = slug

    # Guarantee every uploaded image/table is present with the real data. Runs
    # BEFORE the cap so uploaded images are never counted or stripped.
    _ensure_uploaded_media(
        post,
        job.get("up_images", []),
        job.get("up_tables", []),
        personalized=job.get("personalized", False),
    )

    # Never repeat the same table, image, paragraph, or list within one post.
    _dedupe_content(post)

    # Cap Pexels in-body images. Listicles get one per point; other formats
    # keep 2. Uploaded images (id "upload-img-…") are exempt from the cap.
    img_limit = 10 if job.get("fmt") == "listicle" else 2
    _cap_image_blocks(post, limit=img_limit)

    link_stats = validate_and_clean_links(post, job["allow_primary"], job["allow_secondary"])

    _ensure_internal_links(post, job.get("allow_primary", []))

    base = f"{i:03d}-{slug}"
    post_dir = run_dir / base
    post_dir.mkdir(parents=True, exist_ok=True)

    image_errors: list[str] = []

    # Copy uploaded images into the post dir (these bypass Pexels entirely).
    _copy_uploaded_images(post, slug, post_dir, job.get("up_images", []), image_errors)

    # Collect the remaining image blocks (hero + Pexels in-body) to fetch.
    image_blocks: list[dict] = []
    hero_block = post.get("hero", {}).get("image")
    if hero_block:
        hero_block.setdefault("id", "hero")
        hero_block.setdefault("src", f"/assets/blogs/{slug}/hero.png")
        image_blocks.append({"_target": hero_block, "id": "hero",
                             "alt": hero_block.get("alt", ""), "caption": ""})
    for b in post.get("content", []):
        if b.get("type") == "image" and not str(b.get("id", "")).startswith("upload-img-"):
            image_blocks.append({"_target": b, "id": b.get("id", "image"),
                                 "alt": b.get("alt", ""), "caption": b.get("caption", "")})

    if job["make_images"]:
        for img in image_blocks:
            filename = f"{img['id']}.png"
            target = img["_target"]
            try:
                img_bytes, mime = image_gen.generate_image(_image_prompt(img, slug))
                if "jpeg" in mime:
                    filename = filename.replace(".png", ".jpg")
                (post_dir / filename).write_bytes(img_bytes)
                target["src"] = f"/assets/blogs/{slug}/{filename}"
            except Exception as exc:  # noqa: BLE001
                image_errors.append(f"{img['id']}: {exc}")

    # Normalise into the CMS export schema (data-wrapped tables, image_id on
    # every image, sanitised inline html) so the JSON download matches the
    # format the downstream CMS expects.
    editing.sanitize_post(post)

    # Persist all output formats.
    (post_dir / "post.json").write_text(
        json.dumps(post, indent=2, ensure_ascii=False), encoding="utf-8")
    (post_dir / "post.md").write_text(renderer.render_markdown(post), encoding="utf-8")
    (post_dir / "post.html").write_text(
        renderer.render_html(post, asset_dir=post_dir), encoding="utf-8")
    try:
        renderer.render_pdf(post, post_dir / "post.pdf", asset_dir=post_dir)
    except Exception as exc:  # noqa: BLE001
        image_errors.append(f"pdf: {exc}")
    try:
        renderer.render_docx(post, post_dir / "post.docx", asset_dir=post_dir)
    except Exception as exc:  # noqa: BLE001
        image_errors.append(f"docx: {exc}")

    # Mirror the finished post into durable storage so preview/edit/download
    # survive a server restart (no-op when DATABASE_URL is unset).
    store.save_dir(f"{run_id}/{base}", post_dir)

    meta = post.get("meta", {})
    seo = post.get("seo", {})
    hero_file = (
        f"{run_id}/{base}/hero.png" if (post_dir / "hero.png").exists()
        else f"{run_id}/{base}/hero.jpg" if (post_dir / "hero.jpg").exists()
        else None
    )
    return {
        "index": i, "keyword": keyword, "slug": slug,
        "title": meta.get("title") or seo.get("title", ""),
        "subtitle": meta.get("subtitle", ""),
        "description": seo.get("description", ""),
        "tags": meta.get("tags", []),
        "category": meta.get("category", ""),
        "word_count": _word_count(post),
        "links": link_stats,
        "hero_image": hero_file,
        "downloads": {
            "json":     f"{run_id}/{base}/post.json",
            "markdown": f"{run_id}/{base}/post.md",
            "html":     f"{run_id}/{base}/post.html",
            "pdf":      f"{run_id}/{base}/post.pdf",
            "docx":     f"{run_id}/{base}/post.docx",
        },
        "image_errors": image_errors,
        "usage": written["usage"],
    }


def _ensure_internal_links(post: dict, primary_urls: list[str]) -> None:
    """Guarantee the post has internal links to primary URLs.

    If the model produced fewer than 3 internal links, inject them into the
    CTA paragraph (and body paragraphs as needed) so every blog ships with
    at least 3 Hexa links regardless of what the model decided to do.
    """
    if not primary_urls:
        return

    existing = 0
    for b in post.get("content", []):
        for lnk in b.get("links", []):
            if lnk.get("kind") == "internal":
                existing += 1
    if existing >= 3:
        return

    _ANCHORS = [
        ("Hexa Climate", None),
        ("renewable energy solutions", None),
        ("our renewable energy projects", None),
        ("green energy for businesses", None),
        ("explore our blog", None),
        ("talk to our team", None),
        ("learn more about Hexa", None),
    ]

    main_url = primary_urls[0]
    blogs_url = next((u for u in primary_urls if "/blog" in u.lower()), None)
    projects_url = next((u for u in primary_urls if "/project" in u.lower() or "/re-" in u.lower()), None)

    # Rotation order: CTA gets the main site; body paragraphs cycle through
    # the distinct deep pages so we never link the same URL twice.
    rotation = [u for u in (blogs_url, projects_url, main_url) if u]
    for u in primary_urls:
        if u not in rotation:
            rotation.append(u)

    already_used = set()
    for b in post.get("content", []):
        for lnk in b.get("links", []):
            already_used.add(lnk.get("href"))

    needed = max(0, 3 - existing)
    injected = 0
    content = post.get("content", [])

    for b in reversed(content):
        if injected >= needed:
            break
        if b.get("type") != "paragraph":
            continue
        text = b.get("text", "")
        if not text or len(text) < 40:
            continue
        links = b.setdefault("links", [])
        bid = b.get("id", "")
        is_cta = "cta" in bid.lower()

        if is_cta and main_url not in already_used:
            url = main_url
        else:
            url = next((u for u in rotation if u not in already_used), None)
        if url is None:
            break
        already_used.add(url)

        for anchor_text, _ in _ANCHORS:
            if anchor_text.lower() in text.lower():
                idx = text.lower().index(anchor_text.lower())
                exact = text[idx:idx + len(anchor_text)]
                links.append({"anchor": exact, "href": url, "kind": "internal"})
                injected += 1
                break
        else:
            last_sentence_start = text.rfind(". ")
            if last_sentence_start > 0:
                snippet = text[last_sentence_start + 2:]
            else:
                snippet = text
            words = snippet.split()
            if len(words) >= 3:
                anchor = " ".join(words[:3])
                if anchor in text:
                    links.append({"anchor": anchor, "href": url, "kind": "internal"})
                    injected += 1


def _cap_image_blocks(post: dict, *, limit: int = 2) -> None:
    """Keep at most `limit` Pexels in-body image blocks; strip the rest in place.

    Uploaded images (id "upload-img-…") are user-supplied and always kept.
    """
    content = post.get("content", [])
    seen = 0
    kept: list[dict] = []
    for b in content:
        if b.get("type") == "image" and not str(b.get("id", "")).startswith("upload-img-"):
            if seen >= limit:
                continue
            seen += 1
        kept.append(b)
    post["content"] = kept


# ── Uploaded-media embedding ───────────────────────────────────────────────

def _clean_topic(s: str) -> str:
    """Normalise a candidate topic string (strip extension, separators, noise)."""
    s = re.sub(r"\.[a-z0-9]{1,5}$", "", (s or "").strip(), flags=re.I)  # drop extension
    s = re.sub(r"^(Author's brief for this post:|Blog brief:)\s*", "", s, flags=re.I)
    s = re.sub(r"[_]+", " ", s)               # underscores → spaces
    s = re.sub(r"\s+", " ", s).strip(" -–—:·|")
    return s


def _derive_topic(images: list[dict], tables: list[dict],
                  extra_instructions: str, fallback_keyword: str) -> str:
    """Work out what a personalized post should actually be ABOUT.

    Priority (most resource-specific first):
      1. A description the user typed on an uploaded image/table.
      2. The author's brief (folded into extra_instructions by app.py).
      3. The filename of the first uploaded image/table.
      4. The first row of an uploaded table (its headers).
      5. The queue keyword as a last resort.
    """
    # 1. explicit per-resource descriptions
    for item in list(images) + list(tables):
        d = _clean_topic(item.get("description", ""))
        if len(d) >= 6:
            return d[:120]

    # 2. author's brief (pulled back out of extra_instructions)
    m = re.search(r"Author's brief for this post:\s*(.+)", extra_instructions or "", re.S)
    if m:
        b = _clean_topic(m.group(1))
        if len(b) >= 6:
            return b[:120]

    # 3. filename of the first upload
    for item in list(images) + list(tables):
        n = _clean_topic(item.get("name", ""))
        # skip generic auto names like "image", "img1234", "screenshot"
        if len(n) >= 5 and not re.fullmatch(r"(image|img|photo|screenshot)\s*\d*", n, re.I):
            return n[:120]

    # 4. table header row
    for tb in tables:
        rows = tb.get("rows") or []
        if rows and rows[0]:
            hdr = _clean_topic(" ".join(str(c) for c in rows[0] if c))
            if len(hdr) >= 6:
                return hdr[:120]

    # 5. fall back to whatever keyword was queued
    return (fallback_keyword or "Hexa Climate developments update").strip()


def _media_brief(images: list[dict], tables: list[dict], *, personalized: bool = False) -> str:
    """Tell Claude uploaded media is placed by the system, so it never duplicates it.

    When ``personalized`` is True (Hexa developments run) the tone shifts: the
    uploads become first-priority material, not just decorations — each image
    needs a third-person "Hexa has achieved…" paragraph next to it, and each
    table needs a prose interpretation of its numbers.
    """
    if not images and not tables:
        return ""
    lines: list[str] = [
        "The user uploaded the media below. The SYSTEM inserts each item into the "
        "post automatically, exactly once, with the correct data. Do NOT create "
        "image or table blocks for them yourself, and do NOT restate their data in "
        "another table or list. You may refer to them in prose (for example \"the "
        "table below\" or \"the chart shown\") but never reproduce their contents."
    ]
    if personalized:
        lines.append(
            "PERSONALIZED / HEXA-UPDATE MODE — the focus keyword you were given "
            "already IS the subject of the uploaded resource. Write the entire "
            "post about that subject: title, subtitle, slug, first heading, and "
            "intro paragraph all revolve around it. Do NOT pivot to rooftop "
            "solar pricing, panel prices, or any unrelated consumer topic. "
            "The system places each uploaded image and table near the TOP of "
            "the post (right after the intro paragraph), so write your intro "
            "already introducing what the reader is about to see. For each "
            "image, write a short paragraph describing, in the third person, "
            "what Hexa has achieved / delivered / demonstrated with it "
            "(e.g. \"Hexa has commissioned…\", \"Hexa's team completed…\"). For "
            "each table, add a paragraph that interprets the actual numbers — "
            "biggest movers, year-on-year change, milestone the figures represent — "
            "in the same third-person Hexa voice. Do not treat any upload as "
            "decoration or filler."
        )
    if images:
        lines.append("Uploaded images (placed automatically):")
        for im in images:
            lines.append(f"  - {im.get('description') or 'an uploaded image'}")
    if tables:
        lines.append("Uploaded data tables (placed automatically):")
        for tb in tables:
            lines.append(f"  - {tb.get('description') or 'an uploaded data table'}")
    return "\n".join(lines)


def _insert_before_cta(content: list[dict], block: dict) -> None:
    """Insert `block` just before the final heading (usually the CTA)."""
    for idx in range(len(content) - 1, -1, -1):
        if content[idx].get("type") == "heading":
            content.insert(idx, block)
            return
    content.append(block)


def _insert_after_intro(content: list[dict], block: dict) -> None:
    """Insert `block` high up in the post — right after the first paragraph that
    follows the first heading. This is the placement used in personalized mode
    so uploaded resources become the spine of the post rather than tail decoration.
    """
    seen_heading = False
    for idx, b in enumerate(content):
        t = b.get("type")
        if t == "heading":
            seen_heading = True
            continue
        if seen_heading and t == "paragraph":
            content.insert(idx + 1, block)
            return
    # Fallback — no intro found; insert as the second block so it's still near the top.
    content.insert(min(1, len(content)), block)


def _ensure_uploaded_media(
    post: dict,
    images: list[dict],
    tables: list[dict],
    *,
    personalized: bool = False,
) -> None:
    """Insert each uploaded image/table exactly once, with verified data.

    The pipeline owns placement (Claude is told not to recreate them). Any block
    the model may still have produced for these ids is dropped first so nothing
    is duplicated, and captions never leak the raw filename. In personalized
    mode uploads are placed near the top (the post is about them); otherwise
    they sit just before the CTA.
    """
    slug = post.get("slug", "")
    up_ids = ({f"upload-table-{t['id']}" for t in tables}
              | {f"upload-img-{i['id']}" for i in images})
    content = [b for b in post.get("content", []) if b.get("id") not in up_ids]
    place = _insert_after_intro if personalized else _insert_before_cta

    for tb in tables:
        # Uploaded tables ship with row[0] as headers; wrap in the CMS
        # `data` shape so the export matches the schema.
        rows = tb["rows"] or []
        headers = rows[0] if rows else []
        body = rows[1:] if rows else []
        place(content, {
            "id": f"upload-table-{tb['id']}", "type": "table",
            "data": {"headers": list(headers), "rows": [list(r) for r in body]},
            "caption": (tb.get("description") or "").strip(),
        })
    for im in images:
        place(content, {
            "id": f"upload-img-{im['id']}", "type": "image",
            "src": f"/assets/blogs/{slug}/{im['stored_as']}",
            "alt": (im.get("description") or "Renewable energy visual from Hexa Climate").strip(),
            "caption": (im.get("description") or "").strip(),
        })
    post["content"] = content


# ── De-duplication (no repeated tables, images, paragraphs, or lists) ──────

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _block_signature(block: dict):
    """A hashable identity for a block, or None if it shouldn't be de-duped."""
    t = block.get("type")
    if t == "paragraph":
        s = _norm(block.get("text", ""))
        return ("paragraph", s) if s else None
    if t == "heading":
        s = _norm(block.get("text", ""))
        return ("heading", block.get("level", 2), s) if s else None
    if t == "list":
        items = tuple(_norm(i) for i in block.get("items", []))
        return ("list", items) if any(items) else None
    if t == "image":
        # Dedupe only true repeats (same picture AND same alt), so distinct
        # per-point listicle images that share a placeholder src are all kept.
        key = _norm(block.get("src", "")) + "|" + _norm(block.get("alt", ""))
        return ("image", key) if key.strip("|") else None
    if t == "table":
        data = block.get("data") if isinstance(block.get("data"), dict) else {}
        raw_rows = data.get("rows", block.get("rows", []))
        rows = tuple(tuple(str(c) for c in r) for r in raw_rows)
        return ("table", rows) if rows else None
    if t == "quote":
        s = _norm(block.get("text", ""))
        return ("quote", s) if s else None
    return None


def _dedupe_content(post: dict) -> None:
    """Drop any block whose content repeats one already kept earlier in the post."""
    seen: set = set()
    kept: list[dict] = []
    for block in post.get("content", []):
        sig = _block_signature(block)
        if sig is not None and sig in seen:
            continue
        if sig is not None:
            seen.add(sig)
        kept.append(block)
    post["content"] = kept


def _copy_uploaded_images(post: dict, slug: str, post_dir: Path,
                          images: list[dict], image_errors: list[str]) -> None:
    """Copy each uploaded image file next to the post and point its block at it."""
    by_id = {f"upload-img-{im['id']}": im for im in images}
    for b in post.get("content", []):
        if b.get("type") != "image":
            continue
        im = by_id.get(b.get("id"))
        if not im:
            continue
        src_path = uploads.stored_path(im["stored_as"])
        if src_path.exists():
            shutil.copyfile(src_path, post_dir / im["stored_as"])
            b["src"] = f"/assets/blogs/{slug}/{im['stored_as']}"
        else:
            image_errors.append(f"upload-img-{im['id']}: stored file missing")


_PEXELS_STYLE_HINTS = (
    "renewable energy, clean energy, professional photography, editorial, "
    "sustainability, industry, real-world scene"
)


def _image_prompt(block: dict, slug: str) -> str:
    """Build a Pexels-friendly query — real photographable subject + style hint."""
    alt = (block.get("alt") or "").strip()
    cap = (block.get("caption") or "").strip()
    base = alt or cap or f"renewable energy article about {slug.replace('-', ' ')}"
    # Strip diagram/infographic language so Pexels doesn't match irrelevant schematics.
    base = re.sub(
        r"\b(diagram|infographic|chart|flow(chart)?|schematic|illustration|graphic)\b",
        "", base, flags=re.IGNORECASE,
    ).strip(" ,.")
    return f"{base}. {_PEXELS_STYLE_HINTS}"


def _word_count(post: dict) -> int:
    n = 0
    for b in post.get("content", []):
        if b.get("type") == "paragraph":
            n += len((b.get("text") or "").split())
        elif b.get("type") == "list":
            n += sum(len((i or "").split()) for i in b.get("items", []))
        elif b.get("type") == "heading":
            n += len((b.get("text") or "").split())
    return n
